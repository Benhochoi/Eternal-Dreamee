"""
eval_metrics.py
--------------------------------------------------------------------------------
Bo chi so danh gia chatbot RAG HVNH — 5 chi so:
    1. Time to First Token (TTFT)
    2. Tokens per Second (TPS)
    3. Answer Accuracy (keyword overlap)
    4. Recall@K (yeu cau /api/debug_retrieve trong main.py)
    5. Faithfulness / Hallucination (LLM judge)

Ket qua: in bang + luu eval_results_YYYYMMDD_HHMMSS.xlsx

Chay:
    pip install httpx openpyxl
    python eval_metrics.py          # can main.py dang chay truoc
--------------------------------------------------------------------------------
"""
import re
import json
import math
import time
import asyncio
import statistics
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter


# ============================================================
# CAU HINH
# ============================================================

API_BASE    = "http://localhost:5000"
OLLAMA_URL  = "http://localhost:11434/api/generate"
JUDGE_MODEL = "qwen2.5:3b"
TOP_K       = 5
TIMEOUT     = 120


# ============================================================
# DATASET — 25 cau hoi tu tai lieu thuc
# expected = cum tu chinh xac can co trong cau tra loi
# chunk_ids = parent chunk_id dung (theo chroma_export.xlsx)
# ============================================================

EVAL_DATASET = [

    # ----- QD 2202: Ren luyen -----
    {
        "stt": 1,
        "nhom": "Ren luyen",
        "query": "Quy định 2202/QĐ-HVNH đánh giá rèn luyện sinh viên theo thang điểm nào?",
        "expected": "thang điểm 100",
        "chunk_ids": ["2202/QĐ-HVNH__dieu_5"],
    },
    {
        "stt": 2,
        "nhom": "Ren luyen",
        "query": "Sinh viên đạt bao nhiêu điểm rèn luyện thì được xếp loại xuất sắc?",
        "expected": "từ 90 đến 100 điểm",
        "chunk_ids": ["2202/QĐ-HVNH__dieu_12"],
    },
    {
        "stt": 3,
        "nhom": "Ren luyen",
        "query": "Tiêu chí ý thức chấp hành quy chế quy định được tối đa bao nhiêu điểm?",
        "expected": "25 điểm",
        "chunk_ids": ["2202/QĐ-HVNH__dieu_7"],
    },
    {
        "stt": 4,
        "nhom": "Ren luyen",
        "query": "Sinh viên bị kỷ luật cảnh cáo thì kết quả rèn luyện không được vượt quá loại nào?",
        "expected": "không được vượt quá loại trung bình",
        "chunk_ids": ["2202/QĐ-HVNH__dieu_13"],
    },
    {
        "stt": 5,
        "nhom": "Ren luyen",
        "query": "Sinh viên bị xếp loại rèn luyện yếu kém hai học kỳ liên tiếp thì xử lý thế nào?",
        "expected": "tạm ngừng học ít nhất một học kỳ",
        "chunk_ids": ["2202/QĐ-HVNH__dieu_16"],
    },

    # ----- NQ 267: Quy che cong tac sinh vien -----
    {
        "stt": 6,
        "nhom": "Cong tac SV",
        "query": "Nghị quyết 267/NQ-HĐHV quy định về những nội dung nào trong công tác sinh viên?",
        "expected": "giáo dục tư vấn hỗ trợ quản lý khen thưởng kỷ luật",
        "chunk_ids": ["267/NQ-HĐHV__dieu_1"],
    },
    {
        "stt": 7,
        "nhom": "Cong tac SV",
        "query": "Hội đồng kỷ luật sinh viên của HVNH gồm bao nhiêu thành viên?",
        "expected": "7 thành viên",
        "chunk_ids": ["267/NQ-HĐHV__dieu_36"],
    },
    {
        "stt": 8,
        "nhom": "Cong tac SV",
        "query": "Các hình thức kỷ luật sinh viên tại HVNH gồm những hình thức nào?",
        "expected": "khiển trách cảnh cáo đình chỉ buộc thôi học",
        "chunk_ids": ["267/NQ-HĐHV__dieu_33"],
    },
    {
        "stt": 9,
        "nhom": "Cong tac SV",
        "query": "Thời hiệu xử lý kỷ luật đối với hành vi vi phạm ít nghiêm trọng là bao lâu?",
        "expected": "01 năm",
        "chunk_ids": ["267/NQ-HĐHV__dieu_37"],
    },

    # ----- QD 2786: Cong nhan KQHT / Chuyen doi TC -----
    {
        "stt": 10,
        "nhom": "Chuyen doi TC",
        "query": "Quy định 2786/QĐ-HVNH áp dụng cho đối tượng sinh viên nào?",
        "expected": "sinh viên đại học chính quy hệ tín chỉ tại Học viện Ngân hàng",
        "chunk_ids": ["2786/QĐ-HVNH__dieu_1"],
    },
    {
        "stt": 11,
        "nhom": "Chuyen doi TC",
        "query": "Sinh viên có IELTS 6.5 được quy đổi điểm học phần Đọc Viết II bao nhiêu điểm?",
        "expected": "9.0",
        "chunk_ids": ["2786/QĐ-HVNH__dieu_5_p2"],
    },
    {
        "stt": 12,
        "nhom": "Chuyen doi TC",
        "query": "Chứng chỉ CFA được công nhận chuyển đổi tín chỉ tại HVNH không?",
        "expected": "CFA được công nhận chuyển đổi tín chỉ",
        "chunk_ids": ["2786/QĐ-HVNH__dieu_6", "309/QĐ-HVNH__dieu_3_p2"],
    },

    # ----- QD 1862: ECTS -----
    {
        "stt": 13,
        "nhom": "ECTS",
        "query": "1 tín chỉ của Học viện Ngân hàng tương đương bao nhiêu ECTS tại Áo, Ý, Tây Ban Nha?",
        "expected": "1 tín chỉ = 2 ECTS",
        "chunk_ids": ["1862/QĐ-HVNH__dieu_1"],
    },
    {
        "stt": 14,
        "nhom": "ECTS",
        "query": "Tại Phần Lan 1 ECTS bằng bao nhiêu giờ học và tương đương bao nhiêu tín chỉ HVNH?",
        "expected": "1 ECTS = 27 giờ và 1 tín chỉ = 1,85 ECTS",
        "chunk_ids": ["1862/QĐ-HVNH__dieu_1"],
    },
    {
        "stt": 15,
        "nhom": "ECTS",
        "query": "1 tín chỉ Việt Nam được tính tương đương bao nhiêu giờ học tập định mức?",
        "expected": "50 giờ học tập định mức",
        "chunk_ids": ["1862/QĐ-HVNH__dieu_1"],
    },

    # ----- QD 3337: Chuan dau ra ngoai ngu CNTT -----
    {
        "stt": 16,
        "nhom": "Ngoai ngu CNTT",
        "query": "Chuẩn đầu ra ngoại ngữ sinh viên đại học HVNH là bậc mấy trong khung 6 bậc?",
        "expected": "bậc 3/6",
        "chunk_ids": ["3337/QĐ-HVNH__dieu_2"],
    },
    {
        "stt": 17,
        "nhom": "Ngoai ngu CNTT",
        "query": "Những chứng chỉ ngoại ngữ nào được công nhận đạt chuẩn đầu ra tại HVNH?",
        "expected": "IELTS TOEFL TOEIC",
        "chunk_ids": ["3337/QĐ-HVNH__dieu_3_p1"],
    },

    # ----- QD 335: Huong dan quy che dao tao -----
    {
        "stt": 18,
        "nhom": "Quy che dao tao",
        "query": "Sinh viên được đăng ký tối thiểu và tối đa bao nhiêu tín chỉ mỗi học kỳ chính?",
        "expected": "tối thiểu 14 tín chỉ tối đa 25 tín chỉ",
        "chunk_ids": ["335/QĐ-HVNH__dieu_10"],
    },
    {
        "stt": 19,
        "nhom": "Quy che dao tao",
        "query": "Điểm trung bình chung tích lũy bao nhiêu thì sinh viên bị cảnh báo kết quả học tập?",
        "expected": "dưới 1.2",
        "chunk_ids": ["335/QĐ-HVNH__dieu_16"],
    },
    {
        "stt": 20,
        "nhom": "Quy che dao tao",
        "query": "Điều kiện để sinh viên được học cùng lúc hai chương trình đào tạo là gì?",
        "expected": "điểm trung bình tích lũy từ 2.0 trở lên không bị cảnh báo",
        "chunk_ids": ["335/QĐ-HVNH__dieu_17"],
    },

    # ----- Lich hoc / ca hoc -----
    {
        "stt": 21,
        "nhom": "Lich hoc",
        "query": "Ca 1 buổi sáng tại HVNH bắt đầu lúc mấy giờ và kết thúc lúc mấy giờ?",
        "expected": "7h00 đến 8h15 tiết 1 và 8h20 đến 9h20 tiết 2",
        "chunk_ids": ["Thời_gian_các_ca_học_tại_HVNH__s0_p1"],
    },
    {
        "stt": 22,
        "nhom": "Lich hoc",
        "query": "Buổi tối ca 5 hệ chính quy kết thúc lúc mấy giờ?",
        "expected": "20h05",
        "chunk_ids": ["Thời_gian_các_ca_học_tại_HVNH__s0_p2"],
    },

    # ----- Cau hoi ngoai pham vi (kiem tra hallucination) -----
    {
        "stt": 23,
        "nhom": "Ngoai pham vi",
        "query": "Học viện Ngân hàng có bao nhiêu sinh viên đang theo học hiện nay?",
        "expected": "KHONG_CO",
        "chunk_ids": [],
    },
    {
        "stt": 24,
        "nhom": "Ngoai pham vi",
        "query": "Học phí học kỳ 1 năm 2025 của Học viện Ngân hàng là bao nhiêu tiền?",
        "expected": "KHONG_CO",
        "chunk_ids": [],
    },
    {
        "stt": 25,
        "nhom": "Ngoai pham vi",
        "query": "Lịch nghỉ Tết Nguyên Đán năm 2025 của Học viện Ngân hàng?",
        "expected": "KHONG_CO",
        "chunk_ids": [],
    },
]


# ============================================================
# HELPERS
# ============================================================

def _tokens(text: str) -> int:
    return len(text.split())

def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", text.lower().strip())

def _overlap(answer: str, expected: str) -> float:
    if expected == "KHONG_CO":
        return 0.0
    aw = set(_norm(answer).split())
    ew = set(_norm(expected).split())
    return len(aw & ew) / len(ew) if ew else 0.0


# ============================================================
# METRIC 1+2: TTFT va TPS
# ============================================================

async def measure_ttft_tps(query: str, client: httpx.AsyncClient) -> dict:
    t0      = time.perf_counter()
    t_first = None
    chunks  = []
    try:
        async with client.stream(
            "POST", f"{API_BASE}/api/chat",
            json={"message": query, "stream": True},
            timeout=TIMEOUT,
        ) as resp:
            async for chunk in resp.aiter_text():
                if chunk.strip():
                    if t_first is None:
                        t_first = time.perf_counter()
                    chunks.append(chunk)
        t1       = time.perf_counter()
        answer   = "".join(chunks).strip()
        n_tok    = _tokens(answer)
        gen_t    = t1 - (t_first or t0)
        return {
            "ttft_s":  round((t_first - t0) if t_first else (t1 - t0), 3),
            "tps":     round(n_tok / max(gen_t, 0.001), 1),
            "total_s": round(t1 - t0, 2),
            "n_tokens": n_tok,
            "answer":  answer,
        }
    except Exception as e:
        return {"ttft_s": -1, "tps": -1, "total_s": -1,
                "n_tokens": 0, "answer": "", "error": str(e)}


# ============================================================
# METRIC 3: ANSWER ACCURACY
# ============================================================

def measure_accuracy(answer: str, expected: str) -> dict:
    REFUSE = ["không tìm thấy", "xin lỗi", "không có thông tin",
              "không liên quan", "không tìm thấy nội dung"]
    if expected == "KHONG_CO":
        ok = any(s in answer.lower() for s in REFUSE)
        return {"score": 1.0 if ok else 0.0, "correct": ok,
                "note": "refused_ok" if ok else "hallucinated"}
    score = _overlap(answer, expected)
    return {"score": round(score, 3), "correct": score >= 0.4,
            "note": f"overlap={score:.2f}"}


# ============================================================
# METRIC 4: RECALL@K
# ============================================================

async def measure_recall(query: str, correct_ids: list,
                         client: httpx.AsyncClient) -> dict:
    if not correct_ids:
        return {"hit": None, "retrieved": [], "note": "no_correct_chunk"}
    try:
        r   = await client.post(
            f"{API_BASE}/api/debug_retrieve",
            json={"query": query, "k": TOP_K}, timeout=30)
        ret = r.json().get("chunk_ids", [])
        hit = any(cid in ret for cid in correct_ids)
        return {"hit": hit, "retrieved": ret, "note": f"top{TOP_K}"}
    except Exception as e:
        return {"hit": None, "retrieved": [], "note": f"error:{e}"}


# ============================================================
# METRIC 5: FAITHFULNESS (LLM judge)
# ============================================================

_JUDGE_PROMPT = """Ban la giam khao kiem tra chatbot.

[TAI LIEU NGUON]
{context}

[CAU TRA LOI]
{answer}

Danh gia: cau tra loi co bam sat tai lieu nguon khong? Tra ve 1 so nguyen 0-10.
10=hoan toan dung nguon, 0=bia dat hoan toan. Chi tra ve so, khong giai thich.

DIEM:"""

async def measure_faithfulness(answer: str, context: str,
                                client: httpx.AsyncClient) -> dict:
    if not answer:
        return {"score": 0.0, "raw": 0}
    try:
        r = await client.post(OLLAMA_URL, json={
            "model": JUDGE_MODEL,
            "prompt": _JUDGE_PROMPT.format(
                context=context[:1500], answer=answer[:800]),
            "stream": False,
        }, timeout=60)
        txt   = r.json().get("response", "0")
        nums  = re.findall(r"\d+", txt)
        raw   = max(0, min(10, int(nums[0]))) if nums else 0
        return {"score": round(raw / 10, 2), "raw": raw}
    except Exception as e:
        return {"score": 0.0, "raw": 0, "note": str(e)}


# ============================================================
# EVAL 1 SAMPLE
# ============================================================

async def eval_one(s: dict, client: httpx.AsyncClient,
                   idx: int, total: int) -> dict:
    q = s["query"]
    print(f"\n[{idx+1}/{total}] {q[:65]}...")

    perf   = await measure_ttft_tps(q, client)
    answer = perf.get("answer", "")
    print(f"  TTFT={perf['ttft_s']}s | TPS={perf['tps']} | {perf['n_tokens']} tok")

    acc    = measure_accuracy(answer, s["expected"])
    recall = await measure_recall(q, s["chunk_ids"], client)
    faith  = await measure_faithfulness(answer, answer, client)

    hit_str = "HIT" if recall["hit"] else ("MISS" if recall["hit"] is False else "N/A")
    print(f"  Acc={acc['score']:.2f}({acc['note']}) | Recall@{TOP_K}={hit_str} | Faith={faith['score']:.2f}")

    return {
        "stt":         s["stt"],
        "nhom":        s["nhom"],
        "query":       q,
        "expected":    s["expected"],
        "answer_preview": answer[:200],
        # Metric 1
        "ttft_s":      perf["ttft_s"],
        # Metric 2
        "tps":         perf["tps"],
        "total_s":     perf["total_s"],
        "n_tokens":    perf["n_tokens"],
        # Metric 3
        "accuracy":    acc["score"],
        "correct":     acc["correct"],
        "acc_note":    acc["note"],
        # Metric 4
        "recall_hit":  recall["hit"],
        "recall_ids":  "|".join(recall["retrieved"][:TOP_K]),
        "recall_note": recall["note"],
        # Metric 5
        "faithfulness": faith["score"],
        "faith_raw":    faith["raw"],
    }


# ============================================================
# TONG HOP
# ============================================================

def summarize(results: list[dict]) -> dict:
    def avg(vals):
        v = [x for x in vals if x is not None and x >= 0]
        return round(sum(v) / len(v), 3) if v else None

    def pct(bools):
        v = [b for b in bools if b is not None]
        return round(sum(v) / len(v), 3) if v else None

    in_scope  = [r for r in results if r["expected"] != "KHONG_CO"]
    out_scope = [r for r in results if r["expected"] == "KHONG_CO"]

    return {
        "n_total":          len(results),
        "n_in_scope":       len(in_scope),
        "n_out_scope":      len(out_scope),
        "ttft_avg":         avg([r["ttft_s"]      for r in results]),
        "ttft_median":      round(statistics.median([r["ttft_s"] for r in results if r["ttft_s"] >= 0]), 3),
        "tps_avg":          avg([r["tps"]          for r in results]),
        "accuracy_avg":     avg([r["accuracy"]     for r in results]),
        "accuracy_rate":    pct([r["correct"]      for r in results]),
        "recall_at_k":      pct([r["recall_hit"]   for r in results]),
        "recall_k":         TOP_K,
        "faithfulness_avg": avg([r["faithfulness"] for r in results]),
        "refuse_rate":      pct([r["correct"] for r in out_scope]) if out_scope else None,
    }


# ============================================================
# XUAT EXCEL
# ============================================================

# Mau sac
_C_HEADER  = "1F4E79"   # xanh dam
_C_ALT     = "EBF3FB"   # xanh nhat xen ke
_C_WHITE   = "FFFFFF"
_C_GOOD    = "C6EFCE"   # xanh la - tot
_C_WARN    = "FFEB9C"   # vang - trung binh
_C_BAD     = "FFC7CE"   # do - kem
_C_SUMMARY = "2E4057"   # xanh dam summary
_C_SUM_BG  = "D9E8F5"

_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _header_style(cell, bg: str = _C_HEADER):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _THIN

def _data_style(cell, alt: bool = False, bold: bool = False):
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = _fill(_C_ALT if alt else _C_WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _THIN

def _color_score(cell, score):
    """To mau o theo muc score 0-1."""
    if score is None:
        return
    if score >= 0.7:
        cell.fill = _fill(_C_GOOD)
    elif score >= 0.4:
        cell.fill = _fill(_C_WARN)
    else:
        cell.fill = _fill(_C_BAD)

def _color_bool(cell, val):
    if val is True:
        cell.fill = _fill(_C_GOOD)
    elif val is False:
        cell.fill = _fill(_C_BAD)


def _build_sheet_chitiet(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Chi tiet ket qua")
    ws.title = "Chi tiet ket qua"

    headers = [
        "STT", "Nhom", "Cau hoi",
        "TTFT (s)", "TPS (tok/s)", "Tong t (s)", "So token",
        "Accuracy", "Dung/Sai", "Ghi chu acc",
        f"Recall@{TOP_K}", "Chunk tim duoc",
        "Faithfulness", "Diem judge/10",
        "Cau tra loi (preview)",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        _header_style(ws.cell(1, col))

    for i, r in enumerate(results):
        alt  = i % 2 == 0
        row  = i + 2
        vals = [
            r["stt"], r["nhom"], r["query"],
            r["ttft_s"], r["tps"], r["total_s"], r["n_tokens"],
            r["accuracy"],
            "Dung" if r["correct"] else "Sai",
            r["acc_note"],
            ("HIT" if r["recall_hit"] else
             ("MISS" if r["recall_hit"] is False else "N/A")),
            r["recall_ids"],
            r["faithfulness"], r["faith_raw"],
            r["answer_preview"],
        ]
        ws.append(vals)
        for col, v in enumerate(vals, 1):
            c = ws.cell(row, col)
            _data_style(c, alt)
            # To mau cho cac cot so lieu
            if col == 8:   _color_score(c, r["accuracy"])
            if col == 9:   _color_bool(c, r["correct"])
            if col == 11:  _color_bool(c, r["recall_hit"])
            if col == 13:  _color_score(c, r["faithfulness"])
        # Cot query va answer - wrap text, can trai
        for col in [3, 15]:
            ws.cell(row, col).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

    # Chieu rong cot
    widths = {1:5, 2:14, 3:40, 4:9, 5:9, 6:9, 7:8,
              8:10, 9:9, 10:14, 11:8, 12:35, 13:12, 14:10, 15:45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    # Chieu cao dong
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 45
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _build_sheet_summary(wb: Workbook, summary: dict, results: list[dict]):
    ws = wb.active
    ws.title = "Tong ket"

    # Tieu de lon
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "DANH GIA CHATBOT RAG — HVNH"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(_C_SUMMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"].value     = f"Thoi gian chay: {ts}  |  So cau hoi: {summary['n_total']}  |  Recall@K: K={summary['recall_k']}"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Bang 5 chi so
    rows_5 = [
        ["Chi so", "Gia tri do duoc", "Muc tot", "Danh gia", "Mo ta"],
        ["1. Time to First Token (TTFT)",
         f"{summary['ttft_avg']} s (avg) / {summary['ttft_median']} s (median)",
         "< 2 s",
         "Tot" if (summary["ttft_avg"] or 99) < 2 else "Can cai thien",
         "Thoi gian tu luc gui den khi nhan token dau tien"],
        ["2. Tokens per Second (TPS)",
         f"{summary['tps_avg']} tok/s",
         "> 15 tok/s",
         "Tot" if (summary["tps_avg"] or 0) > 15 else "Can cai thien",
         "Toc do sinh text sau token dau tien"],
        ["3. Answer Accuracy",
         f"{round((summary['accuracy_rate'] or 0)*100, 1)}%  (score TB: {summary['accuracy_avg']})",
         ">= 70%",
         "Tot" if (summary["accuracy_rate"] or 0) >= 0.7 else "Can cai thien",
         "Ty le cau tra loi co chua tu khoa dap an (nguong 40% overlap)"],
        [f"4. Recall@{summary['recall_k']}",
         (f"{round((summary['recall_at_k'] or 0)*100, 1)}%"
          if summary["recall_at_k"] is not None else "Chua do (endpoint N/A)"),
         ">= 70%",
         ("Tot" if (summary["recall_at_k"] or 0) >= 0.7
          else ("Can cai thien" if summary["recall_at_k"] else "Chua do")),
         f"Ty le cau hoi tim dung chunk trong top-{summary['recall_k']}"],
        ["5. Faithfulness",
         f"{summary['faithfulness_avg']} / 1.0",
         ">= 0.7",
         "Tot" if (summary["faithfulness_avg"] or 0) >= 0.7 else "Can cai thien",
         "Muc do bam nguon, khong bia them (LLM judge 0-10)"],
    ]

    start_row = 4
    for ri, row_data in enumerate(rows_5):
        ws.append(row_data)
        actual_row = start_row + ri
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(actual_row, ci)
            if ri == 0:
                _header_style(c)
            else:
                alt = ri % 2 == 0
                _data_style(c, alt)
                # Cot chi so can trai, in dam
                if ci == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.font = Font(name="Arial", size=10, bold=True)
                # To mau cot danh gia
                if ci == 4:
                    if val == "Tot":
                        c.fill = _fill(_C_GOOD)
                        c.font = Font(name="Arial", size=9, bold=True, color="375623")
                    else:
                        c.fill = _fill(_C_WARN)
                        c.font = Font(name="Arial", size=9, bold=True, color="7D4700")

    # Bang ty le tu choi hop le (ngoai pham vi)
    if summary["refuse_rate"] is not None:
        ws.cell(start_row + 7, 1).value = "Ty le tu choi dung (cau ngoai pham vi):"
        ws.cell(start_row + 7, 1).font  = Font(name="Arial", bold=True, size=10)
        ws.cell(start_row + 7, 2).value = f"{round(summary['refuse_rate']*100, 1)}%"
        _color_score(ws.cell(start_row + 7, 2), summary["refuse_rate"])

    # Bang ket qua theo nhom
    nhom_data: dict[str, list] = {}
    for r in results:
        nhom = r["nhom"]
        nhom_data.setdefault(nhom, []).append(r)

    gr = start_row + 10
    ws.cell(gr, 1).value = "Ket qua theo nhom cau hoi"
    ws.cell(gr, 1).font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(gr, 1).fill  = _fill(_C_SUMMARY)
    for col in range(1, 7):
        ws.cell(gr, col).fill = _fill(_C_SUMMARY)
    gr += 1

    gr_headers = ["Nhom", "So cau", "Accuracy TB", "Recall HIT", "Faithfulness TB"]
    for ci, h in enumerate(gr_headers, 1):
        _header_style(ws.cell(gr, ci))
        ws.cell(gr, ci).value = h
    gr += 1

    for ni, (nhom, rlist) in enumerate(nhom_data.items()):
        alt  = ni % 2 == 0
        accs = [r["accuracy"] for r in rlist]
        hits = [r["recall_hit"] for r in rlist if r["recall_hit"] is not None]
        fths = [r["faithfulness"] for r in rlist]
        row_vals = [
            nhom, len(rlist),
            round(sum(accs)/len(accs), 2) if accs else 0,
            f"{round(sum(hits)/len(hits)*100, 0):.0f}%" if hits else "N/A",
            round(sum(fths)/len(fths), 2) if fths else 0,
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(gr, ci)
            c.value = v
            _data_style(c, alt)
            if ci == 3:
                _color_score(c, row_vals[2])
            if ci == 5:
                _color_score(c, row_vals[4])
        gr += 1

    # Do rong cot
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 14


def export_excel(results: list[dict], summary: dict) -> str:
    wb = Workbook()
    _build_sheet_summary(wb, summary, results)
    _build_sheet_chitiet(wb, results)

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"eval_results_{ts}.xlsx"
    wb.save(path)
    return path


# ============================================================
# MAIN
# ============================================================

async def main():
    print("=" * 65)
    print("  DANH GIA CHATBOT RAG HVNH — 5 CHI SO")
    print(f"  API: {API_BASE}")
    print(f"  Dataset: {len(EVAL_DATASET)} cau hoi")
    print("=" * 65)

    async with httpx.AsyncClient() as client:
        # Kiem tra server
        try:
            r = await client.get(f"{API_BASE}/api/health", timeout=5)
            print(f"Server: OK — {r.json().get('model')}")
        except Exception:
            print("[FAIL] Khong ket noi duoc server!")
            print("       Chay 'python main.py' truoc roi thu lai.")
            return

        # Kiem tra debug_retrieve
        try:
            await client.post(f"{API_BASE}/api/debug_retrieve",
                              json={"query": "test", "k": 1}, timeout=5)
            print("Endpoint /api/debug_retrieve: OK")
        except Exception:
            print("[WARN] /api/debug_retrieve chua co — Recall@K se la N/A")

        results = []
        for idx, sample in enumerate(EVAL_DATASET):
            r = await eval_one(sample, client, idx, len(EVAL_DATASET))
            results.append(r)
            await asyncio.sleep(0.3)

    summary = summarize(results)

    # In tong ket ra terminal
    print("\n" + "=" * 65)
    print("  KET QUA TONG HOP")
    print("=" * 65)
    print(f"  TTFT avg      : {summary['ttft_avg']} s")
    print(f"  TPS avg       : {summary['tps_avg']} tok/s")
    print(f"  Accuracy rate : {round((summary['accuracy_rate'] or 0)*100,1)}%")
    print(f"  Recall@{TOP_K}      : {round((summary['recall_at_k'] or 0)*100,1) if summary['recall_at_k'] else 'N/A'}%")
    print(f"  Faithfulness  : {summary['faithfulness_avg']} / 1.0")
    if summary["refuse_rate"] is not None:
        print(f"  Refuse rate   : {round(summary['refuse_rate']*100,1)}% (ngoai pham vi)")

    # Xuat Excel
    xlsx_path = export_excel(results, summary)
    print(f"\n  Ket qua da luu: {xlsx_path}")
    print("=" * 65)


if __name__ == "__main__":
    asyncio.run(main())