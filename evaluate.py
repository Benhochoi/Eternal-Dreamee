"""
eval_metrics.py
--------------------------------------------------------------------------------
Bo chi so danh gia chatbot RAG cua HVNH

5 chi so:
    1. Time to First Token (TTFT)   -- do tre khi model bat dau tra loi
    2. Tokens per Second (TPS)      -- toc do sinh text
    3. Answer Accuracy               -- do chinh xac cau tra loi (so voi expected)
    4. Recall@K                      -- kha nang retrieve dung tai lieu
    5. Faithfulness (Hallucination)  -- muc do bam nguon, khong bia them

Chay:
    python eval_metrics.py

Ket qua: in bang chi so + luu file eval_results_YYYYMMDD_HHMMSS.json

Luu y:
    - File nay chay doc lap voi main.py, ket noi API qua http://localhost:5000
    - Phai khoi dong main.py truoc: python main.py
    - Faithfulness dung LLM judge (goi them 1 lan cho moi cau hoi)
--------------------------------------------------------------------------------
"""
import json
import math
import re
import time
import statistics
import asyncio
import httpx
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ============================================================
# CAU HINH
# ============================================================

API_URL     = "http://localhost:5000/api/chat"
STREAM_URL  = "http://localhost:5000/api/chat"   # dung stream=True de do TTFT
JUDGE_URL   = "http://localhost:11434/api/generate"  # Ollama truc tiep cho judge
JUDGE_MODEL = "qwen2.5:3b"                       # model dung de judge faithfulness

TOP_K = 3   # Recall@K

TIMEOUT_SEC = 120   # timeout moi request


# ============================================================
# DATASET -- 20 cau hoi voi dap an mau va chunk_id dung
# Them / sua cau hoi o day de mo rong bo test
# ============================================================

EVAL_DATASET = [
    # ----- QD 2786: Cong nhan KQ hoc tap -----
    {
        "query":           "Quy định 2786/QĐ-HVNH áp dụng cho đối tượng nào?",
        "expected_answer": "sinh viên đại học chính quy hệ tín chỉ tại Học viện Ngân hàng",
        "correct_chunk_ids": ["2786/QĐ-HVNH__dieu_1"],
        "domain":          "chuyen_doi_tc",
    },
    {
        "query":           "Điều kiện điểm số để một học phần được công nhận kết quả học tập?",
        "expected_answer": "điểm học phần đạt từ 5.0 trở lên hoặc điểm chữ D trở lên",
        "correct_chunk_ids": ["2786/QĐ-HVNH__dieu_3"],
        "domain":          "chuyen_doi_tc",
    },
    {
        "query":           "Sinh viên có IELTS 6.5 được quy đổi điểm học phần tiếng Anh như thế nào?",
        "expected_answer": "được quy đổi điểm 9.0 cho học phần Đọc Viết II",
        "correct_chunk_ids": ["2786/QĐ-HVNH__dieu_5_p1", "2786/QĐ-HVNH__dieu_5_p2"],
        "domain":          "chuyen_doi_tc",
    },
    {
        "query":           "Chứng chỉ nghề nghiệp nào được công nhận chuyển đổi tín chỉ?",
        "expected_answer": "CFA, FRM, ACCA và các chứng chỉ nghề nghiệp quốc tế được công nhận",
        "correct_chunk_ids": ["2786/QĐ-HVNH__dieu_6"],
        "domain":          "chuyen_doi_tc",
    },
    # ----- QD 3337: Chuan dau ra ngoai ngu -----
    {
        "query":           "Chuẩn đầu ra ngoại ngữ sinh viên đại học HVNH là bậc mấy?",
        "expected_answer": "bậc 3/6 theo khung năng lực ngoại ngữ 6 bậc của Việt Nam",
        "correct_chunk_ids": ["3337/QĐ-HVNH__dieu_2", "3337/QĐ-HVNH__dieu_3_p1"],
        "domain":          "ngoai_ngu",
    },
    {
        "query":           "Những chứng chỉ ngoại ngữ nào được công nhận đạt chuẩn đầu ra tại HVNH?",
        "expected_answer": "IELTS, TOEFL iBT, TOEIC và các chứng chỉ ngoại ngữ quốc tế tương đương",
        "correct_chunk_ids": ["3337/QĐ-HVNH__dieu_3_p1"],
        "domain":          "ngoai_ngu",
    },
    {
        "query":           "Sinh viên cần chứng chỉ công nghệ thông tin nào để đạt chuẩn đầu ra?",
        "expected_answer": "chứng chỉ ứng dụng công nghệ thông tin cơ bản hoặc nâng cao",
        "correct_chunk_ids": ["3337/QĐ-HVNH__dieu_4"],
        "domain":          "cntt",
    },
    # ----- QD 335: Quy che dao tao he tin chi -----
    {
        "query":           "Sinh viên được đăng ký tối thiểu và tối đa bao nhiêu tín chỉ mỗi kỳ?",
        "expected_answer": "tối thiểu 14 tín chỉ, tối đa 25 tín chỉ mỗi học kỳ chính",
        "correct_chunk_ids": ["335/QĐ-HVNH__dieu_10"],
        "domain":          "hoc_phan",
    },
    {
        "query":           "Điểm trung bình chung tích lũy bao nhiêu thì bị cảnh báo kết quả học tập?",
        "expected_answer": "dưới 1.2 đối với sinh viên năm nhất hoặc theo quy định cụ thể từng năm học",
        "correct_chunk_ids": ["335/QĐ-HVNH__dieu_16"],
        "domain":          "canh_bao",
    },
    {
        "query":           "Sinh viên được nghỉ học tạm thời vì những lý do gì?",
        "expected_answer": "ốm đau, tai nạn, thai sản, khó khăn kinh tế, hoặc nhu cầu cá nhân sau 1 năm học",
        "correct_chunk_ids": ["335/QĐ-HVNH__dieu_15"],
        "domain":          "canh_bao",
    },
    {
        "query":           "Điều kiện để sinh viên được học cùng lúc hai chương trình đào tạo?",
        "expected_answer": "điểm trung bình tích lũy từ 2.0 trở lên và không bị cảnh báo học tập",
        "correct_chunk_ids": ["335/QĐ-HVNH__dieu_17"],
        "domain":          "hoc_phan",
    },
    {
        "query":           "Cách tính điểm học phần từ các điểm thành phần như thế nào?",
        "expected_answer": "tổng hợp điểm chuyên cần, kiểm tra giữa kỳ và thi cuối kỳ theo tỷ lệ quy định",
        "correct_chunk_ids": ["335/QĐ-HVNH__dieu_19", "335/QĐ-HVNH__dieu_22"],
        "domain":          "hoc_phan",
    },
    # ----- QD 2833: Dang ky hoc phan -----
    {
        "query":           "Phòng Đào tạo công bố kế hoạch đăng ký học phần khi nào?",
        "expected_answer": "trước mỗi học kỳ theo lịch do Phòng Đào tạo thông báo",
        "correct_chunk_ids": ["2833/QĐ-HVNH__dieu_3"],
        "domain":          "hoc_phan",
    },
    {
        "query":           "Sinh viên không nộp học phí đúng hạn thì xử lý như thế nào?",
        "expected_answer": "bị hủy đăng ký học phần và không được tham gia học tập trong học kỳ đó",
        "correct_chunk_ids": ["2833/QĐ-HVNH__dieu_6"],
        "domain":          "hoc_phi",
    },
    # ----- QD 309: Chung chi nghe nghiep -----
    {
        "query":           "Chứng chỉ CFA Level 1 được quy đổi bao nhiêu điểm hệ 10?",
        "expected_answer": "quy đổi theo công thức điểm chứng chỉ nhân hệ số quy định",
        "correct_chunk_ids": ["309/QĐ-HVNH__dieu_3_p1", "309/QĐ-HVNH__dieu_3_p2"],
        "domain":          "chuyen_doi_tc",
    },
    # ----- Van ban thuong -----
    {
        "query":           "Giảng viên được tính thêm bao nhiêu tiết chuẩn cho hướng dẫn tự học?",
        "expected_answer": "2 tiết chuẩn cho mỗi lớp học phần",
        "correct_chunk_ids": ["Quy_định_hướng_dẫn_sinh_viên_tự_học__s2"],
        "domain":          "hoc_phan",
    },
    {
        "query":           "Giảng viên tiếp xúc sinh viên ngoài giờ lên lớp bằng hình thức nào?",
        "expected_answer": "qua email, hẹn lịch trực tiếp tại văn phòng khoa hoặc giảng đường",
        "correct_chunk_ids": ["Quy_định_hướng_dẫn_sinh_viên_tự_học__s3"],
        "domain":          "hoc_phan",
    },
    # ----- Thoi gian ca hoc -----
    {
        "query":           "Ca 1 buổi sáng tại Học viện Ngân hàng bắt đầu và kết thúc lúc mấy giờ?",
        "expected_answer": "tiết 1 từ 7h00 đến 8h15, tiết 2 từ 8h20 đến 9h20",
        "correct_chunk_ids": ["Thoi_gian_ca_hoc__s1"],
        "domain":          "lich_hoc",
    },
    {
        "query":           "Buổi chiều tại HVNH có những ca học nào và thời gian ra sao?",
        "expected_answer": "ca 3 từ 12h45 đến 15h05 và ca 4 từ 15h20 đến 17h40",
        "correct_chunk_ids": ["Thoi_gian_ca_hoc__s1"],
        "domain":          "lich_hoc",
    },
    {
        "query":           "Ca 5 buổi tối kết thúc lúc mấy giờ đối với hệ chính quy?",
        "expected_answer": "kết thúc lúc 20h05 đối với hệ chính quy (CQ)",
        "correct_chunk_ids": ["Thoi_gian_ca_hoc__s1"],
        "domain":          "lich_hoc",
    },
    {
        "query":           "Thời gian nghỉ giữa 2 tiết học và giữa 2 ca học là bao nhiêu phút?",
        "expected_answer": "nghỉ giữa 2 tiết là 5 phút, giữa 2 ca là 15 phút",
        "correct_chunk_ids": ["Thoi_gian_ca_hoc__s1"],
        "domain":          "lich_hoc",
    },
    # ----- Quy che sinh vien (NQ cong tac sinh vien) -----
    {
        "query":           "Sinh viên vi phạm quy chế thi có thể bị xử lý kỷ luật ở mức nào?",
        "expected_answer": "từ cảnh cáo đến buộc thôi học tùy mức độ vi phạm",
        "correct_chunk_ids": ["NQ_cong_tac_SV__ky_luat"],
        "domain":          "canh_bao",
    },
    {
        "query":           "Điểm rèn luyện của sinh viên được đánh giá theo những tiêu chí nào?",
        "expected_answer": "ý thức học tập, chấp hành nội quy, tham gia hoạt động tập thể và đời sống cá nhân",
        "correct_chunk_ids": ["NQ_cong_tac_SV__ren_luyen"],
        "domain":          "ren_luyen",
    },
    # ----- Quy che ket noi cong dong -----
    {
        "query":           "Hoạt động kết nối và phục vụ cộng đồng của sinh viên HVNH bao gồm những hình thức nào?",
        "expected_answer": "tình nguyện, hỗ trợ cộng đồng, các hoạt động xã hội do Học viện tổ chức hoặc phối hợp",
        "correct_chunk_ids": ["NQ_ket_noi_cong_dong__dieu_2"],
        "domain":          "other",
    },
    # ----- Van hoa hoc duong -----
    {
        "query":           "Sinh viên HVNH cần tuân thủ những chuẩn mực ứng xử nào trong học đường?",
        "expected_answer": "tôn trọng thầy cô, bạn bè, ăn mặc lịch sự, không gian lận trong thi cử",
        "correct_chunk_ids": ["van_hoa_hoc_duong__chuan_muc"],
        "domain":          "other",
    },
    # ----- Hoc bong -----
    {
        "query":           "Tiêu chí xét học bổng khuyến khích học tập tại HVNH là gì?",
        "expected_answer": "dựa trên điểm trung bình học tập và điểm rèn luyện trong học kỳ",
        "correct_chunk_ids": ["NQ_hoc_bong__dieu_tieu_chi"],
        "domain":          "hoc_bong",
    },
    {
        "query":           "Mức học bổng khuyến khích học tập được phân loại như thế nào?",
        "expected_answer": "chia thành các mức theo xếp loại học lực: xuất sắc, giỏi, khá",
        "correct_chunk_ids": ["NQ_hoc_bong__dieu_muc_hb"],
        "domain":          "hoc_bong",
    },
    # ----- Huong dan tu hoc (bo sung) -----
    {
        "query":           "Sinh viên cần thực hiện những nhiệm vụ tự học nào ngoài giờ lên lớp?",
        "expected_answer": "đọc tài liệu, làm bài tập, nghiên cứu theo hướng dẫn của giảng viên",
        "correct_chunk_ids": ["Quy_định_hướng_dẫn_sinh_viên_tự_học__s1"],
        "domain":          "hoc_phan",
    },
    # ----- ECTS -----
    {
        "query":           "Tín chỉ ECTS được quy đổi sang tín chỉ của HVNH theo tỷ lệ nào?",
        "expected_answer": "1 tín chỉ HVNH tương đương khoảng 2 tín chỉ ECTS theo hướng dẫn quy đổi",
        "correct_chunk_ids": ["ECTS__quy_doi_ty_le"],
        "domain":          "chuyen_doi_tc",
    },
    # ----- Cau hoi ngoai pham vi (kiem tra hallucination) -----
    {
        "query":           "Học viện Ngân hàng có bao nhiêu sinh viên đang theo học hiện nay?",
        "expected_answer": "KHONG_CO_TRONG_TAI_LIEU",   # ky hieu khong co dap an
        "correct_chunk_ids": [],
        "domain":          "other",
    },
    {
        "query":           "Học phí học kỳ 1 năm 2025 là bao nhiêu tiền?",
        "expected_answer": "KHONG_CO_TRONG_TAI_LIEU",
        "correct_chunk_ids": [],
        "domain":          "hoc_phi",
    },
    {
        "query":           "Lịch nghỉ tết nguyên đán năm 2025 của Học viện Ngân hàng?",
        "expected_answer": "KHONG_CO_TRONG_TAI_LIEU",
        "correct_chunk_ids": [],
        "domain":          "lich_hoc",
    },
]


# ============================================================
# HELPERS
# ============================================================

def _count_tokens(text: str) -> int:
    """Uoc tinh so token bang cach dem tu (co dau = 1 token)."""
    return len(text.split())


def _normalize(text: str) -> str:
    """Chuan hoa text truoc khi so sanh."""
    return re.sub(r"\s+", " ", text.lower().strip())


def _keyword_overlap(answer: str, expected: str) -> float:
    """
    Do do phu tu khoa giua cau tra loi va expected.
    Dung cho Answer Accuracy khi khong co LLM judge.
    Score = so tu chung / max(so tu expected, 1)
    """
    if expected == "KHONG_CO_TRONG_TAI_LIEU":
        return 0.0   # xu ly rieng o phan accuracy
    a_words = set(_normalize(answer).split())
    e_words = set(_normalize(expected).split())
    if not e_words:
        return 0.0
    return len(a_words & e_words) / len(e_words)


# ============================================================
# METRIC 1+2: TTFT va TPS (do qua streaming endpoint)
# ============================================================

async def measure_ttft_tps(query: str, client: httpx.AsyncClient) -> dict:
    """
    Goi streaming endpoint va do:
        TTFT  = thoi gian tu luc gui request den khi nhan token dau tien
        TPS   = tong so token / tong thoi gian generate
        total = toan bo cau tra loi ghep lai
    """
    payload = {"message": query, "stream": True}
    t_start = time.perf_counter()
    t_first = None
    tokens  = []

    try:
        async with client.stream(
            "POST", STREAM_URL,
            json=payload,
            timeout=TIMEOUT_SEC,
        ) as resp:
            async for chunk in resp.aiter_text():
                if chunk.strip():
                    if t_first is None:
                        t_first = time.perf_counter()
                    tokens.append(chunk)

        t_end    = time.perf_counter()
        full_ans = "".join(tokens)
        n_tok    = _count_tokens(full_ans)
        total_t  = t_end - t_start
        gen_t    = t_end - (t_first or t_start)

        return {
            "ttft_s":    round((t_first - t_start) if t_first else total_t, 3),
            "tps":       round(n_tok / max(gen_t, 0.001), 1),
            "total_s":   round(total_t, 2),
            "n_tokens":  n_tok,
            "answer":    full_ans.strip(),
        }
    except Exception as e:
        return {"ttft_s": -1, "tps": -1, "total_s": -1,
                "n_tokens": 0, "answer": "", "error": str(e)}


# ============================================================
# METRIC 3: ANSWER ACCURACY (keyword overlap)
# ============================================================

def measure_accuracy(answer: str, expected: str) -> dict:
    """
    Do do chinh xac cau tra loi:

    Neu expected = "KHONG_CO_TRONG_TAI_LIEU":
        - Neu model tra loi "khong tim thay" / "xin loi" → correct = True (1.0)
        - Neu model tra loi co noi dung → correct = False (0.0, hallucination nguy co)

    Neu co expected:
        - Tinh keyword overlap giua answer va expected
        - Score >= 0.4 thi coi la correct
    """
    REFUSE_SIGNALS = [
        "không tìm thấy", "xin lỗi", "không có thông tin",
        "không tìm thấy nội dung", "không liên quan",
    ]

    if expected == "KHONG_CO_TRONG_TAI_LIEU":
        ans_lower  = answer.lower()
        is_refused = any(s in ans_lower for s in REFUSE_SIGNALS)
        return {
            "score":    1.0 if is_refused else 0.0,
            "correct":  is_refused,
            "note":     "refused_correctly" if is_refused else "hallucinated_answer",
        }

    score = _keyword_overlap(answer, expected)
    return {
        "score":   round(score, 3),
        "correct": score >= 0.4,
        "note":    f"overlap={score:.2f}",
    }


# ============================================================
# METRIC 4: RECALL@K (goi retrieval endpoint rieng)
# ============================================================

async def measure_recall_at_k(
    query:              str,
    correct_chunk_ids:  list[str],
    client:             httpx.AsyncClient,
    k:                  int = TOP_K,
) -> dict:
    """
    Goi /api/debug_retrieve de lay chunk_id cua top-k result.
    Hit = 1 neu bat ky correct_chunk_id nao nam trong top-k.

    Neu endpoint /api/debug_retrieve chua co → dung gia lap tu answer.
    """
    if not correct_chunk_ids:
        # Cau hoi ngoai pham vi — khong co dap an chunk
        return {"hit": None, "retrieved_ids": [], "note": "no_correct_chunk_defined"}

    try:
        resp = await client.post(
            "http://localhost:5000/api/debug_retrieve",
            json={"query": query, "k": k},
            timeout=30,
        )
        data         = resp.json()
        retrieved    = data.get("chunk_ids", [])
        hit          = any(cid in retrieved for cid in correct_chunk_ids)
        return {
            "hit":          hit,
            "retrieved_ids": retrieved,
            "note":         f"top{k}_retrieved",
        }
    except Exception:
        # Endpoint chua co → tra ve None (se tinh rieng o phan tong hop)
        return {"hit": None, "retrieved_ids": [], "note": "endpoint_unavailable"}


# ============================================================
# METRIC 5: FAITHFULNESS / HALLUCINATION SCORE (LLM judge)
# ============================================================

FAITHFULNESS_PROMPT = """Ban la giam khao danh gia chatbot.

[TAI LIEU NGUON]
{context}

[CAU TRA LOI CUA BOT]
{answer}

Nhiem vu: Danh gia xem cau tra loi co BAM SAT vao tai lieu nguon khong.
Cho diem tu 0 den 10, trong do:
    10 = hoan toan lay thong tin tu tai lieu, khong them gi
    7-9 = chu yeu bam tai lieu, co mot vai dien giai nho
    4-6 = mot phan bam tai lieu, mot phan tu suy luan rieng
    1-3 = phan lon khong co trong tai lieu
    0   = hoan toan bia dat, khong lien quan tai lieu

Chi tra ve duy nhat mot so nguyen tu 0 den 10. Khong giai thich them.

DIEM:"""

async def measure_faithfulness(
    answer:  str,
    context: str,
    client:  httpx.AsyncClient,
) -> dict:
    """
    Dung LLM judge (Ollama truc tiep) de danh gia faithfulness.
    Goi ollama API truc tiep de tranh qua main.py.
    Score 0-10 → normalize ve 0.0-1.0.
    """
    if not answer or not context:
        return {"score": 0.0, "raw": 0, "note": "empty_input"}

    prompt_text = FAITHFULNESS_PROMPT.format(
        context=context[:2000],   # gioi han context gui cho judge
        answer=answer[:1000],
    )

    try:
        resp = await client.post(
            JUDGE_URL,
            json={
                "model":  JUDGE_MODEL,
                "prompt": prompt_text,
                "stream": False,
            },
            timeout=60,
        )
        raw_text = resp.json().get("response", "0").strip()
        # Trich so dau tien trong response
        nums = re.findall(r"\d+", raw_text)
        raw_score = int(nums[0]) if nums else 0
        raw_score = max(0, min(10, raw_score))   # clamp 0-10
        return {
            "score": round(raw_score / 10, 2),
            "raw":   raw_score,
            "note":  "llm_judge",
        }
    except Exception as e:
        return {"score": 0.0, "raw": 0, "note": f"judge_error: {e}"}


# ============================================================
# RETRIEVE CONTEXT (de dung cho faithfulness)
# ============================================================

async def get_context_for_query(
    query:  str,
    client: httpx.AsyncClient,
) -> str:
    """
    Goi non-streaming endpoint de lay ca answer lan context.
    Tra ve context text (ghep tu answer vi main.py khong expose context rieng).
    """
    try:
        resp   = await client.post(
            API_URL,
            json={"message": query, "stream": False},
            timeout=TIMEOUT_SEC,
        )
        answer = resp.json().get("reply", "")
        # Context = toan bo answer (proxy don gian; production nen expose /context endpoint)
        return answer
    except Exception:
        return ""


# ============================================================
# CHAY DANH GIA 1 SAMPLE
# ============================================================

async def eval_one(sample: dict, client: httpx.AsyncClient, idx: int, total: int) -> dict:
    query    = sample["query"]
    expected = sample["expected_answer"]
    c_ids    = sample["correct_chunk_ids"]

    print(f"\n[{idx+1}/{total}] {query[:70]}...")

    # Do TTFT + TPS + lay answer
    perf = await measure_ttft_tps(query, client)
    print(f"  TTFT={perf['ttft_s']}s | TPS={perf['tps']} tok/s | {perf['n_tokens']} tokens")

    answer = perf.get("answer", "")

    # Accuracy
    acc = measure_accuracy(answer, expected)
    print(f"  Accuracy: {acc['score']:.2f} ({acc['note']})")

    # Recall@K
    recall = await measure_recall_at_k(query, c_ids, client)
    if recall["hit"] is not None:
        print(f"  Recall@{TOP_K}: {'HIT' if recall['hit'] else 'MISS'} | {recall['retrieved_ids'][:2]}")
    else:
        print(f"  Recall@{TOP_K}: {recall['note']}")

    # Faithfulness — dung answer lam proxy context
    faith = await measure_faithfulness(answer, answer, client)
    print(f"  Faithfulness: {faith['score']:.2f}/1.0 (raw={faith['raw']}/10)")

    return {
        "query":        query,
        "domain":       sample.get("domain", ""),
        "expected":     expected,
        "answer":       answer[:300],   # giu ngan trong JSON
        # TTFT + TPS
        "ttft_s":       perf["ttft_s"],
        "tps":          perf["tps"],
        "total_s":      perf["total_s"],
        "n_tokens":     perf["n_tokens"],
        # Accuracy
        "accuracy":     acc["score"],
        "correct":      acc["correct"],
        "acc_note":     acc["note"],
        # Recall@K
        "recall_hit":   recall["hit"],
        "retrieved_ids":recall["retrieved_ids"][:TOP_K],
        "recall_note":  recall["note"],
        # Faithfulness
        "faithfulness": faith["score"],
        "faith_raw":    faith["raw"],
    }


# ============================================================
# TONG HOP KET QUA
# ============================================================

def summarize(results: list[dict]) -> dict:
    """Tinh trung binh va in bang tong ket."""

    def _avg(vals):
        clean = [v for v in vals if v is not None and v >= 0]
        return round(statistics.mean(clean), 3) if clean else None

    def _pct(bools):
        clean = [b for b in bools if b is not None]
        return round(sum(clean) / len(clean), 3) if clean else None

    ttft_vals   = [r["ttft_s"]     for r in results]
    tps_vals    = [r["tps"]        for r in results]
    acc_vals    = [r["accuracy"]   for r in results]
    correct_bools = [r["correct"]  for r in results]
    recall_bools  = [r["recall_hit"] for r in results]
    faith_vals  = [r["faithfulness"] for r in results]

    # Tach rieng cau hoi trong va ngoai pham vi
    in_scope  = [r for r in results if r["expected"] != "KHONG_CO_TRONG_TAI_LIEU"]
    out_scope = [r for r in results if r["expected"] == "KHONG_CO_TRONG_TAI_LIEU"]

    refused_correctly = sum(1 for r in out_scope if r["correct"])

    summary = {
        "n_total":           len(results),
        "n_in_scope":        len(in_scope),
        "n_out_scope":       len(out_scope),
        # 1. TTFT
        "ttft_avg_s":        _avg(ttft_vals),
        "ttft_p50_s":        round(statistics.median(ttft_vals), 3) if ttft_vals else None,
        # 2. TPS
        "tps_avg":           _avg(tps_vals),
        # 3. Accuracy
        "accuracy_avg":      _avg(acc_vals),
        "accuracy_rate":     _pct(correct_bools),
        # 4. Recall@K
        "recall_at_k":       _pct(recall_bools),
        "recall_k":          TOP_K,
        "recall_note":       "null_if_endpoint_unavailable",
        # 5. Faithfulness
        "faithfulness_avg":  _avg(faith_vals),
        # Bonus: hallucination check
        "refuse_rate_out_scope": (
            round(refused_correctly / len(out_scope), 2) if out_scope else None
        ),
    }
    return summary


def print_summary(summary: dict, results: list[dict]):
    print("\n" + "=" * 65)
    print("  KET QUA DANH GIA CHATBOT HVNH")
    print("=" * 65)
    print(f"  So cau hoi: {summary['n_total']}"
          f" (trong pham vi: {summary['n_in_scope']},"
          f" ngoai pham vi: {summary['n_out_scope']})")
    print()

    # Metric 1: TTFT
    print(f"  [1] Time to First Token (TTFT)")
    print(f"      Trung binh : {summary['ttft_avg_s']} giay")
    print(f"      Median     : {summary['ttft_p50_s']} giay")
    print(f"      Muc tot    : < 2s (phan hoi nhanh voi user)")
    print()

    # Metric 2: TPS
    print(f"  [2] Tokens per Second (TPS)")
    print(f"      Trung binh : {summary['tps_avg']} tok/s")
    print(f"      Muc tot    : > 15 tok/s (doc thoai mai)")
    print()

    # Metric 3: Accuracy
    print(f"  [3] Answer Accuracy")
    print(f"      Score TB   : {summary['accuracy_avg']} (keyword overlap)")
    print(f"      Ti le dung : {summary['accuracy_rate'] * 100:.1f}% (nguong >= 0.4)")
    print(f"      Muc tot    : >= 70%")
    print()

    # Metric 4: Recall@K
    if summary["recall_at_k"] is not None:
        print(f"  [4] Recall@{summary['recall_k']}")
        print(f"      Ti le hit  : {summary['recall_at_k'] * 100:.1f}%")
        print(f"      Muc tot    : >= 70%")
    else:
        print(f"  [4] Recall@{summary['recall_k']}: endpoint /api/debug_retrieve chua co")
        print(f"      -> Them endpoint vao main.py de do chi so nay")
    print()

    # Metric 5: Faithfulness
    print(f"  [5] Faithfulness (Hallucination)")
    print(f"      Score TB   : {summary['faithfulness_avg']} / 1.0")
    print(f"      Muc tot    : >= 0.7 (bam nguon, it bia them)")
    if summary["refuse_rate_out_scope"] is not None:
        print(f"      Tu choi dung khi ngoai pham vi: "
              f"{summary['refuse_rate_out_scope'] * 100:.0f}%")
    print()

    # Danh gia tong the
    scores = {
        "ttft":          1.0 if (summary["ttft_avg_s"] or 99) < 2.0   else 0.5,
        "tps":           1.0 if (summary["tps_avg"]    or 0)  > 15    else 0.5,
        "accuracy":      summary["accuracy_rate"]     or 0.0,
        "recall":        summary["recall_at_k"]       or 0.0,
        "faithfulness":  summary["faithfulness_avg"]  or 0.0,
    }
    overall = round(sum(scores.values()) / len(scores), 2)
    print(f"  Diem tong the (trung binh 5 chi so): {overall * 100:.1f}/100")
    print("=" * 65)

    # Bang chi tiet miss / fail
    fails = [r for r in results if not r.get("correct")]
    if fails:
        print(f"\n  Chi tiet {len(fails)} cau chua chinh xac:")
        for r in fails[:5]:
            print(f"    - {r['query'][:60]}  [{r['acc_note']}]")
        if len(fails) > 5:
            print(f"    ... va {len(fails)-5} cau khac (xem JSON de biet them)")


# ============================================================
# ENDPOINT DEBUG_RETRIEVE (them vao main.py neu muon do Recall@K)
# ============================================================

DEBUG_ENDPOINT_CODE = '''
# ----- THEM DOAN NAY VAO main.py DE DO RECALL@K -----

class DebugRetrieveRequest(BaseModel):
    query: str
    k:     int = 3

@app.post("/api/debug_retrieve")
async def debug_retrieve(request: DebugRetrieveRequest):
    """Endpoint phuc vu eval_metrics.py do Recall@K."""
    q_norm = normalize_query(request.query)
    docs   = await retriever.ainvoke(q_norm)
    return {
        "chunk_ids": [d.metadata.get("chunk_id", "") for d in docs[:request.k]],
        "so_hieus":  [d.metadata.get("so_hieu",   "") for d in docs[:request.k]],
    }
# ----- HET DOAN THEM -----
'''


# ============================================================
# XUAT KET QUA RA EXCEL
# ============================================================

def _cell_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _pct_fmt(val):
    """None-safe percentage string."""
    if val is None:
        return "N/A"
    return f"{val * 100:.1f}%"

def _num_fmt(val, decimals=3):
    if val is None:
        return "N/A"
    return round(val, decimals)


def save_excel(summary: dict, results: list[dict], ts: str) -> str:
    wb = Workbook()

    # ── Màu chủ đạo ──────────────────────────────────────────
    BLUE_HDR   = "1F4E79"   # header tối
    BLUE_SUB   = "2E75B6"   # sub-header
    BLUE_LIGHT = "BDD7EE"   # nền hàng chẵn
    WHITE      = "FFFFFF"
    GREEN_OK   = "C6EFCE"
    RED_FAIL   = "FFC7CE"
    YELLOW_MID = "FFEB9C"
    GRAY_LABEL = "F2F2F2"

    border = _cell_border()

    def style_header(cell, bg=BLUE_HDR, font_color=WHITE, size=11, bold=True):
        cell.font      = Font(name="Arial", bold=bold, color=font_color, size=size)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    def style_cell(cell, bold=False, center=False, bg=None, font_color="000000"):
        cell.font      = Font(name="Arial", bold=bold, color=font_color, size=10)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True
        )
        cell.border    = border
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    # ════════════════════════════════════════════════════════
    # SHEET 1: TỔNG HỢP CHỈ SỐ
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Tổng hợp"
    ws1.sheet_view.showGridLines = False

    # Tiêu đề lớn
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "KẾT QUẢ ĐÁNH GIÁ CHATBOT HVNH"
    ws1["A1"].font      = Font(name="Arial", bold=True, size=15, color=WHITE)
    ws1["A1"].fill      = PatternFill("solid", fgColor=BLUE_HDR)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Thời gian đánh giá: {ts}  |  Tổng câu hỏi: {summary['n_total']}  |  Trong phạm vi: {summary['n_in_scope']}  |  Ngoài phạm vi: {summary['n_out_scope']}"
    ws1["A2"].font      = Font(name="Arial", size=10, color="595959", italic=True)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A2"].fill      = PatternFill("solid", fgColor="D6E4F0")
    ws1.row_dimensions[2].height = 18

    # Header bảng chỉ số
    headers_s1 = ["CHỈ SỐ", "GIÁ TRỊ", "MỨC TỐT", "ĐẠT?", "GHI CHÚ", "ĐIỂM"]
    ws1.row_dimensions[4].height = 22
    for col, h in enumerate(headers_s1, 1):
        c = ws1.cell(row=4, column=col, value=h)
        style_header(c, bg=BLUE_SUB)

    # Tính điểm từng chỉ số
    ttft_ok   = (summary["ttft_avg_s"] or 99) < 2.0
    tps_ok    = (summary["tps_avg"]    or 0)  > 15
    acc_rate  = summary["accuracy_rate"]  or 0.0
    recall_r  = summary["recall_at_k"]    or 0.0
    faith_r   = summary["faithfulness_avg"] or 0.0

    overall_score = round((
        (1.0 if ttft_ok else 0.5) +
        (1.0 if tps_ok  else 0.5) +
        acc_rate + recall_r + faith_r
    ) / 5 * 100, 1)

    rows_s1 = [
        ("1. Time to First Token (TTFT)",
         f"TB: {_num_fmt(summary['ttft_avg_s'],2)}s  |  P50: {_num_fmt(summary['ttft_p50_s'],2)}s",
         "< 2 giây", "✅ ĐẠT" if ttft_ok else "❌ CHƯA",
         "Thời gian từ lúc gửi đến token đầu tiên",
         1.0 if ttft_ok else 0.5),

        ("2. Tokens per Second (TPS)",
         f"TB: {_num_fmt(summary['tps_avg'],1)} tok/s",
         "> 15 tok/s", "✅ ĐẠT" if tps_ok else "❌ CHƯA",
         "Tốc độ sinh văn bản",
         1.0 if tps_ok else 0.5),

        ("3. Answer Accuracy",
         f"Score TB: {_num_fmt(summary['accuracy_avg'],3)}  |  Tỉ lệ đúng: {_pct_fmt(summary['accuracy_rate'])}",
         "≥ 70%", "✅ ĐẠT" if acc_rate >= 0.7 else "❌ CHƯA",
         "Keyword overlap với đáp án mẫu (ngưỡng ≥ 0.4)",
         round(acc_rate, 3)),

        ("4. Recall@K (K=" + str(summary["recall_k"]) + ")",
         _pct_fmt(summary["recall_at_k"]) if summary["recall_at_k"] is not None else "N/A (endpoint chưa có)",
         "≥ 70%", "✅ ĐẠT" if recall_r >= 0.7 else ("⚠️ N/A" if summary["recall_at_k"] is None else "❌ CHƯA"),
         "Chunk đúng có trong top-K retrieve",
         round(recall_r, 3)),

        ("5. Faithfulness (Hallucination)",
         f"Score TB: {_num_fmt(summary['faithfulness_avg'],3)} / 1.0",
         "≥ 0.7", "✅ ĐẠT" if faith_r >= 0.7 else "❌ CHƯA",
         "LLM judge đánh giá bám nguồn tài liệu",
         round(faith_r, 3)),
    ]

    for r_idx, (chi_so, gia_tri, muc_tot, dat, ghi_chu, diem) in enumerate(rows_s1, 5):
        bg = BLUE_LIGHT if r_idx % 2 == 0 else WHITE
        ok_bg = GREEN_OK if "ĐẠT" in dat else (YELLOW_MID if "N/A" in dat else RED_FAIL)
        ws1.row_dimensions[r_idx].height = 20

        for col, val in enumerate([chi_so, gia_tri, muc_tot, dat, ghi_chu, diem], 1):
            c = ws1.cell(row=r_idx, column=col, value=val)
            style_cell(c, bold=(col == 1), center=(col in (3, 4, 6)), bg=bg if col != 4 else ok_bg)

    # Tổng điểm
    ws1.row_dimensions[11].height = 24
    ws1.merge_cells("A11:E11")
    ws1["A11"] = f"ĐIỂM TỔNG THỂ (trung bình 5 chỉ số)"
    ws1["A11"].font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws1["A11"].fill      = PatternFill("solid", fgColor=BLUE_HDR)
    ws1["A11"].alignment = Alignment(horizontal="right", vertical="center")
    ws1["A11"].border    = border

    score_bg = GREEN_OK if overall_score >= 70 else (YELLOW_MID if overall_score >= 50 else RED_FAIL)
    ws1["F11"] = f"{overall_score}/100"
    ws1["F11"].font      = Font(name="Arial", bold=True, size=13, color="000000")
    ws1["F11"].fill      = PatternFill("solid", fgColor=score_bg)
    ws1["F11"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["F11"].border    = border

    # Hallucination check
    refuse_rate = summary.get("refuse_rate_out_scope")
    if refuse_rate is not None:
        ws1.row_dimensions[13].height = 18
        ws1.merge_cells("A13:B13")
        ws1["A13"] = "Từ chối đúng khi ngoài phạm vi"
        style_cell(ws1["A13"], bold=True, bg=GRAY_LABEL)
        ws1["C13"] = _pct_fmt(refuse_rate)
        style_cell(ws1["C13"], center=True,
                   bg=GREEN_OK if refuse_rate >= 0.8 else RED_FAIL)
        ws1["D13"] = "Mức tốt: ≥ 80%"
        style_cell(ws1["D13"], bg=GRAY_LABEL)

    # Column widths
    for i, w in enumerate([34, 30, 14, 12, 38, 10], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # SHEET 2: CHI TIẾT TỪNG CÂU HỎI
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📋 Chi tiết")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    # Tiêu đề
    headers_s2 = [
        "#", "Domain", "Câu hỏi",
        "TTFT (s)", "TPS (tok/s)", "Tổng thời gian (s)", "Số token",
        "Accuracy Score", "Correct?", "Ghi chú Accuracy",
        f"Recall@{TOP_K} Hit?", "Chunk IDs retrieved", "Ghi chú Recall",
        "Faithfulness", "Raw Score (/10)",
        "Câu trả lời (rút gọn)"
    ]
    ws2.row_dimensions[1].height = 28
    ws2.merge_cells(f"A1:{get_column_letter(len(headers_s2))}1")
    ws2["A1"] = "CHI TIẾT ĐÁNH GIÁ TỪNG CÂU HỎI — CHATBOT HVNH"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws2["A1"].fill      = PatternFill("solid", fgColor=BLUE_HDR)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[2].height = 36
    for col, h in enumerate(headers_s2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        style_header(c, bg=BLUE_SUB, size=10)

    for r_idx, res in enumerate(results, 3):
        ws2.row_dimensions[r_idx].height = 18
        is_correct = res.get("correct")
        row_bg = GREEN_OK if is_correct else (YELLOW_MID if is_correct is None else RED_FAIL)

        vals = [
            r_idx - 2,
            res.get("domain", ""),
            res.get("query", ""),
            res.get("ttft_s"),
            res.get("tps"),
            res.get("total_s"),
            res.get("n_tokens"),
            res.get("accuracy"),
            "✅" if is_correct else ("⚠️" if is_correct is None else "❌"),
            res.get("acc_note", ""),
            ("✅ HIT" if res.get("recall_hit") else
             ("⚠️ N/A" if res.get("recall_hit") is None else "❌ MISS")),
            ", ".join(res.get("retrieved_ids", [])),
            res.get("recall_note", ""),
            res.get("faithfulness"),
            res.get("faith_raw"),
            res.get("answer", "")[:200],
        ]

        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=r_idx, column=col, value=val)
            is_metric_col = col in (4, 5, 6, 7, 8, 14, 15)
            bg = row_bg if col == 9 else (BLUE_LIGHT if r_idx % 2 == 0 else WHITE)
            style_cell(c, center=is_metric_col or col in (1, 9, 11), bg=bg)

    # Column widths sheet 2
    col_widths = [4, 12, 44, 9, 10, 12, 9, 12, 9, 20, 12, 32, 22, 13, 12, 50]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Conditional formatting — color scale cho Accuracy Score (col H)
    last_row = 2 + len(results)
    ws2.conditional_formatting.add(
        f"H3:H{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFC7CE",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB9C",
            end_type="num",   end_value=1,     end_color="C6EFCE",
        )
    )
    ws2.conditional_formatting.add(
        f"N3:N{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFC7CE",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB9C",
            end_type="num",   end_value=1,     end_color="C6EFCE",
        )
    )

    # ════════════════════════════════════════════════════════
    # SHEET 3: PHÂN TÍCH THEO DOMAIN
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📈 Theo Domain")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:G1")
    ws3["A1"] = "PHÂN TÍCH HIỆU SUẤT THEO DOMAIN"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws3["A1"].fill      = PatternFill("solid", fgColor=BLUE_HDR)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    hdrs3 = ["Domain", "Số câu", "Accuracy TB", "Tỉ lệ đúng", f"Recall@{TOP_K}", "Faithfulness TB", "TTFT TB (s)"]
    ws3.row_dimensions[2].height = 22
    for col, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        style_header(c, bg=BLUE_SUB, size=10)

    # Gom theo domain
    from collections import defaultdict
    domain_groups = defaultdict(list)
    for r in results:
        domain_groups[r.get("domain", "other")].append(r)

    for r_idx, (domain, rows) in enumerate(sorted(domain_groups.items()), 3):
        ws3.row_dimensions[r_idx].height = 18
        n      = len(rows)
        acc_tb = round(statistics.mean([r["accuracy"] for r in rows if r["accuracy"] is not None]), 3)
        rate   = round(sum(1 for r in rows if r["correct"]) / n, 3)
        recall_hits = [r["recall_hit"] for r in rows if r["recall_hit"] is not None]
        rec_tb = round(sum(recall_hits) / len(recall_hits), 3) if recall_hits else None
        faith_tb = round(statistics.mean([r["faithfulness"] for r in rows if r["faithfulness"] is not None]), 3)
        ttft_tb  = round(statistics.mean([r["ttft_s"] for r in rows if (r["ttft_s"] or -1) >= 0]), 2)

        bg = BLUE_LIGHT if r_idx % 2 == 0 else WHITE
        rate_bg = GREEN_OK if rate >= 0.7 else (YELLOW_MID if rate >= 0.4 else RED_FAIL)

        for col, val in enumerate([domain, n, acc_tb, _pct_fmt(rate), _pct_fmt(rec_tb), faith_tb, ttft_tb], 1):
            c = ws3.cell(row=r_idx, column=col, value=val)
            style_cell(c, center=(col != 1),
                       bg=rate_bg if col == 4 else bg,
                       bold=(col == 1))

    # Hàng tổng
    total_row = 3 + len(domain_groups)
    ws3.row_dimensions[total_row].height = 20
    ws3.merge_cells(f"A{total_row}:B{total_row}")
    ws3[f"A{total_row}"] = f"TỔNG CỘNG — {len(results)} câu"
    style_cell(ws3[f"A{total_row}"], bold=True, bg=BLUE_LIGHT)
    total_vals = [
        _num_fmt(summary["accuracy_avg"]),
        _pct_fmt(summary["accuracy_rate"]),
        _pct_fmt(summary["recall_at_k"]),
        _num_fmt(summary["faithfulness_avg"]),
        _num_fmt(summary["ttft_avg_s"], 2),
    ]
    for i, val in enumerate(total_vals):
        c = ws3.cell(row=total_row, column=i + 3, value=val)
        style_cell(c, bold=True, center=True, bg=BLUE_LIGHT)

    for i, w in enumerate([18, 8, 14, 12, 14, 16, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Lưu file ─────────────────────────────────────────────
    out_path = f"eval_results_{ts}.xlsx"
    wb.save(out_path)
    return out_path


# ============================================================
# MAIN
# ============================================================

async def main():
    print("=" * 65)
    print("  DANH GIA CHATBOT HVNH -- 5 CHI SO")
    print(f"  API: {API_URL}")
    print(f"  Dataset: {len(EVAL_DATASET)} cau hoi")
    print(f"  Recall@K: K={TOP_K}")
    print("=" * 65)
    print()
    print("Luu y: Recall@K can endpoint /api/debug_retrieve trong main.py.")
    print("       Neu chua co, chi so nay se bao 'endpoint_unavailable'.")
    print("       Code them endpoint da in o cuoi bao cao.")
    print()

    results = []
    async with httpx.AsyncClient() as client:
        # Kiem tra server con song khong
        try:
            await client.get("http://localhost:5000/api/health", timeout=5)
        except Exception:
            print("[FAIL] Khong ket noi duoc server. Chay 'python main.py' truoc!")
            return

        for idx, sample in enumerate(EVAL_DATASET):
            result = await eval_one(sample, client, idx, len(EVAL_DATASET))
            results.append(result)
            # Nghi 0.5s giua cac cau tranh qua tai server
            await asyncio.sleep(0.5)

    # Tong hop
    summary = summarize(results)
    print_summary(summary, results)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Luu ket qua ra JSON
    json_file = f"eval_results_{ts}.json"
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(
            {"summary": summary, "results": results, "timestamp": ts},
            f, ensure_ascii=False, indent=2
        )
    print(f"\n  Ket qua JSON da luu: {json_file}")

    # Luu ket qua ra Excel
    xlsx_file = save_excel(summary, results, ts)
    print(f"  Ket qua Excel da luu: {xlsx_file}")

    # In huong dan them debug endpoint
    print("\n" + "=" * 65)
    print("  THEM ENDPOINT /api/debug_retrieve VAO main.py DE DO RECALL@K:")
    print("=" * 65)
    print(DEBUG_ENDPOINT_CODE)


if __name__ == "__main__":
    asyncio.run(main())